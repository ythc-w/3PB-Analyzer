"""
Analysis module, containing the core logic for data analysis.
"""

import os
import re
import logging
import statistics
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from sklearn.metrics import r2_score
from sklearn.linear_model import LinearRegression

from utils import get_resource_path
from config import (
    DEFAULT_X_COLUMN, DEFAULT_Y_COLUMN, OUTPUT_IMAGE_DIR, DEFAULT_PRELOAD,
    DEFAULT_Yield_Force_Constant, DEFAULT_Displacement_Constant, Excel_Type,
    DEFAULT_PRELOAD_METHOD, PRELOAD_CONFIRM_POINTS, PRELOAD_TREND_POINTS,
    PRELOAD_MIN_TREND_SLOPE, PRELOAD_BASELINE_POINTS,
    PRELOAD_BASELINE_ABS_TOLERANCE_N, PRELOAD_MIN_POINTS_AFTER,
    PRELOAD_ZERO_DISPLACEMENT, PRELOAD_CORRECT_FORCE_BASELINE,
    STIFFNESS_MIN_POSITIVE_SLOPE, STIFFNESS_MIN_FORCE_SPAN_N,
    STIFFNESS_MIN_FORCE_SPAN_FRACTION, STIFFNESS_MIN_MONOTONIC_FRACTION,
    POSTPEAK_FORCE_FRACTION, POSTPEAK_GRADUAL_MIN_POINTS,
    QC_MAX_EXPECTED_FORCE_N, QC_MAX_EXPECTED_DISPLACEMENT_MM,
)
from preload import correct_baseline, detect_preload, preload_quality_control


def analyse_data_legacy(csv_file, x_column=DEFAULT_X_COLUMN, y_column=DEFAULT_Y_COLUMN, min_window_size=10, max_window_size=20, preload=DEFAULT_PRELOAD,
                        YFC=None, dispc=None):
    """
    Performs linear regression analysis.

    Args:
        csv_file (str): The path to the CSV file.
        x_column (str, optional): The column name for the X-axis data. Defaults to `DEFAULT_X_COLUMN`.
        y_column (str, optional): The column name for the Y-axis data. Defaults to `DEFAULT_Y_COLUMN`.
        min_window_size (int, optional): The minimum size of the linear regression window. Defaults to 10.
        max_window_size (int, optional): The maximum size of the linear regression window. Defaults to 20.
        preload (float, optional): The preload value. Defaults to `DEFAULT_PRELOAD`.

     Returns:
        dict or None: A dictionary containing the analysis results and data for plotting, or None if an error occurs.
    """
    results = {}
    results_plot = {}

    try:
        df = pd.read_csv(csv_file)
        if x_column not in df.columns or y_column not in df.columns:
            raise ValueError(f"Column '{x_column}' or '{y_column}' not found in the CSV file.")

        force_col = df[y_column]
        max_index = force_col.idxmax()
        pre_index_series = force_col[:max_index][force_col[:max_index] <= 0]
        if pre_index_series.empty:
            pre_index = 0
        else:
            pre_index = pre_index_series.idxmax()
        first_index = force_col[force_col.index >= pre_index][force_col[force_col.index >= pre_index] >= preload].index.min()
        if first_index is not None:
            df = df.loc[first_index:]
        else:
            df = pd.DataFrame(columns=df.columns)

        x_data = df[x_column].values
        y_data = df[y_column].values
        max_disp = df[x_column].max()
        max_value = df[y_column].max()
        results["Max Force"] = max_value

        for index_select in range(len(y_data)):
            if y_data[index_select] == max_value:
                mv_index = index_select
        for i in range(mv_index, len(y_data)):
            if y_data[i] < (0.5*max_value):
                x_data = x_data[:i]
                y_data = y_data[:i]
                break

        results_plot["x_data"] = x_data
        results_plot["y_data"] = y_data

        best_r2 = -1
        best_model = None
        best_start = None
        best_end = None

        for current_window_size in range(min_window_size, max_window_size + 1):
            for i in range(0, len(x_data) - current_window_size + 1):
                x_window = x_data[i:i + current_window_size].reshape(-1, 1)
                y_window = y_data[i:i + current_window_size]

                model = LinearRegression()
                model.fit(x_window, y_window)
                y_pred = model.predict(x_window)

                r2 = r2_score(y_window, y_pred)

                if r2 > best_r2:
                    best_r2 = r2
                    best_model = model
                    best_start = i
                    best_end = i + current_window_size

        a = best_model.coef_[0]
        b = best_model.intercept_

        results_plot["a"] = a
        results_plot["b"] = b

        results["Stiffness"] = a

        results_plot["best_start"] = best_start
        results_plot["best_end"] = best_end

        next_x = None
        next_y = None

        if y_data[best_end - 1] == max_value:
            next_x = x_data[best_end - 1]
            next_y = y_data[best_end - 1]
        else:
            if best_end < len(x_data):
                for i in range(best_end, len(x_data)):
                    shifted_expected_y = YFC * a * (x_data[i] - max_disp*dispc) + b
                    if 0.8 * shifted_expected_y <= y_data[i] <= shifted_expected_y:
                        next_x = x_data[i]
                        next_y = y_data[i]
                        break
                    if i == len(x_data) - 1 and next_x is None:
                        next_x = x_data[best_end - 1]
                        next_y = y_data[best_end - 1]

        results["Yield force"] = next_y
        results_plot["yield_force_x"] = next_x
        results_plot["yield_force_y"] = next_y
        postyield_displacement = x_data[-1] - next_x
        results["Postyield Displacement"] = postyield_displacement
        auc = np.trapz(y_data[:], x_data[:])
        results["Work to fracture"] = auc
        output_result = [results["Max Force"], results["Stiffness"], results.get("Yield force", "N/A"),
                         results.get("Postyield Displacement", "N/A"), results["Work to fracture"]]

        return {"results": output_result, "results_plot": results_plot}

    except Exception as e:
        logging.error(f"Error in create_scatter_plot: {e}")
        return None


def _legacy_diagnostics(csv_file, x_column, y_column, preload):
    """Describe the legacy start point without changing legacy calculations."""
    try:
        df = pd.read_csv(csv_file)
        force_col = df[y_column]
        max_index = force_col.idxmax()
        pre_index_series = force_col[:max_index][force_col[:max_index] <= 0]
        pre_index = 0 if pre_index_series.empty else pre_index_series.idxmax()
        eligible = force_col[force_col.index >= pre_index]
        first_index = eligible[eligible >= preload].index.min()
        found = first_index is not None and not pd.isna(first_index)
        return {
            "preload_method": "legacy",
            "raw_points": len(df),
            "threshold": preload,
            "preload_found": found,
            "preload_index": int(first_index) if found else None,
            "preload_force": float(df.loc[first_index, y_column]) if found else None,
            "preload_displacement": float(df.loc[first_index, x_column]) if found else None,
            "legacy_pre_index": int(pre_index),
            "zero_reset": False,
            "qc_status": "WARNING",
            "qc_reasons": ["Legacy single-point preload detection was requested."],
        }
    except Exception as exc:
        return {
            "preload_method": "legacy",
            "preload_found": False,
            "preload_index": None,
            "qc_status": "FAIL",
            "qc_reasons": [f"Legacy preload diagnostics failed: {exc}"],
        }


def _find_best_stiffness_model(x_data, y_data, peak_relative_index,
                               min_window_size, max_window_size,
                               baseline_mad, max_force):
    """Find the legacy highest-R2 window after rejecting non-loading windows."""
    best_r2 = -1
    best_model = None
    best_start = None
    best_end = None
    minimum_span = max(STIFFNESS_MIN_FORCE_SPAN_N,
                       STIFFNESS_MIN_FORCE_SPAN_FRACTION * max_force)
    decrease_tolerance = max(PRELOAD_BASELINE_ABS_TOLERANCE_N,
                             6.0 * (baseline_mad or 0.0))
    fit_limit = peak_relative_index + 1

    for current_window_size in range(min_window_size, max_window_size + 1):
        for i in range(0, fit_limit - current_window_size + 1):
            x_window = x_data[i:i + current_window_size].reshape(-1, 1)
            y_window = y_data[i:i + current_window_size]
            if np.ptp(x_window) <= 0 or np.ptp(y_window) < minimum_span:
                continue
            monotonic_fraction = float(np.mean(np.diff(y_window) >= -decrease_tolerance))
            if monotonic_fraction < STIFFNESS_MIN_MONOTONIC_FRACTION:
                continue

            model = LinearRegression()
            model.fit(x_window, y_window)
            if model.coef_[0] <= STIFFNESS_MIN_POSITIVE_SLOPE:
                continue
            r2 = r2_score(y_window, model.predict(x_window))
            if np.isfinite(r2) and r2 > best_r2:
                best_r2 = r2
                best_model = model
                best_start = i
                best_end = i + current_window_size

    if best_model is None:
        raise ValueError(
            "No positive, pre-peak stiffness window passed the configured QC constraints."
        )
    return best_model, best_start, best_end, best_r2


def analyse_data_robust(csv_file, x_column=DEFAULT_X_COLUMN,
                        y_column=DEFAULT_Y_COLUMN, min_window_size=10,
                        max_window_size=20, preload=DEFAULT_PRELOAD,
                        YFC=None, dispc=None):
    """Analyse a CSV using robust preload detection and traceable QC."""
    try:
        df = pd.read_csv(csv_file)
        if x_column not in df.columns or y_column not in df.columns:
            raise ValueError(
                f"Column '{x_column}' or '{y_column}' not found in the CSV file."
            )
        if min_window_size < 2 or max_window_size < min_window_size:
            raise ValueError("Invalid stiffness window size range.")

        raw_x = pd.to_numeric(df[x_column], errors="coerce").to_numpy(dtype=float)
        raw_y = pd.to_numeric(df[y_column], errors="coerce").to_numpy(dtype=float)
        detection = detect_preload(
            raw_y,
            raw_x,
            preload,
            confirm_points=PRELOAD_CONFIRM_POINTS,
            trend_points=PRELOAD_TREND_POINTS,
            min_trend_slope=PRELOAD_MIN_TREND_SLOPE,
            baseline_points=PRELOAD_BASELINE_POINTS,
            baseline_abs_tolerance=PRELOAD_BASELINE_ABS_TOLERANCE_N,
        )
        detection = preload_quality_control(
            detection,
            raw_y,
            raw_x,
            min_points_after_preload=PRELOAD_MIN_POINTS_AFTER,
            max_expected_force=QC_MAX_EXPECTED_FORCE_N,
            max_expected_displacement=QC_MAX_EXPECTED_DISPLACEMENT_MM,
        )
        if not detection.found or detection.index is None:
            raise ValueError("; ".join(detection.qc_reasons))

        corrected_y = correct_baseline(
            raw_y,
            detection.baseline_force,
            enabled=PRELOAD_CORRECT_FORCE_BASELINE,
        )
        corrected_x = raw_x.copy()
        if PRELOAD_ZERO_DISPLACEMENT:
            corrected_x -= detection.displacement
            detection.zero_reset = True

        source_indices = np.arange(len(df))
        valid_after = (
            (source_indices >= detection.index)
            & np.isfinite(corrected_x)
            & np.isfinite(corrected_y)
        )
        retained_indices = source_indices[valid_after]
        x_trimmed = corrected_x[valid_after]
        y_trimmed = corrected_y[valid_after]
        if len(x_trimmed) < min_window_size:
            raise ValueError("Insufficient finite data after robust preload trimming.")

        max_disp = float(np.max(x_trimmed))
        peak_relative = int(np.argmax(y_trimmed))
        max_value = float(y_trimmed[peak_relative])
        below_half = np.flatnonzero(
            y_trimmed[peak_relative:] < POSTPEAK_FORCE_FRACTION * max_value
        )
        fracture_cut_relative = (
            peak_relative + int(below_half[0]) if len(below_half) else None
        )
        analysis_end = (
            fracture_cut_relative if fracture_cut_relative is not None else len(x_trimmed)
        )
        x_data = x_trimmed[:analysis_end]
        y_data = y_trimmed[:analysis_end]
        analysis_source_indices = retained_indices[:analysis_end]
        if len(x_data) < min_window_size or peak_relative >= len(x_data):
            raise ValueError("Fracture trimming left insufficient analysis data.")

        best_model, best_start, best_end, best_r2 = _find_best_stiffness_model(
            x_data,
            y_data,
            peak_relative,
            min_window_size,
            max_window_size,
            detection.baseline_mad,
            max_value,
        )
        a = float(best_model.coef_[0])
        b = float(best_model.intercept_)

        yfc = DEFAULT_Yield_Force_Constant if YFC is None else YFC
        displacement_constant = (
            DEFAULT_Displacement_Constant if dispc is None else dispc
        )
        next_x = None
        next_y = None
        if y_data[best_end - 1] == max_value:
            next_x = float(x_data[best_end - 1])
            next_y = float(y_data[best_end - 1])
        else:
            for i in range(best_end, len(x_data)):
                shifted_expected_y = (
                    yfc * a * (x_data[i] - max_disp * displacement_constant) + b
                )
                if 0.8 * shifted_expected_y <= y_data[i] <= shifted_expected_y:
                    next_x = float(x_data[i])
                    next_y = float(y_data[i])
                    break
        if next_x is None:
            next_x = float(x_data[best_end - 1])
            next_y = float(y_data[best_end - 1])

        postyield_displacement = float(x_data[-1] - next_x)
        auc = float(np.trapezoid(y_data, x_data))

        qc_reasons = list(detection.qc_reasons)
        qc_status = detection.qc_status

        def add_warning(message):
            nonlocal qc_status
            if qc_status == "PASS":
                qc_status = "WARNING"
            qc_reasons.append(message)

        postpeak_points = analysis_end - peak_relative
        if fracture_cut_relative is None:
            add_warning(
                "Post-peak force never fell below 50% of maximum; work is to the "
                "recorded end, not a confirmed fracture endpoint."
            )
        elif postpeak_points >= POSTPEAK_GRADUAL_MIN_POINTS:
            add_warning(
                f"Post-peak response remained above 50% maximum for "
                f"{postpeak_points - 1} points before the fracture cutoff."
            )

        postpeak_values = y_data[peak_relative:]
        if len(postpeak_values) >= 3:
            rising_fraction = float(np.mean(np.diff(postpeak_values) > 0))
            if rising_fraction > 0.3:
                add_warning(
                    "Post-peak force is substantially non-monotonic; inspect the "
                    "fracture tail manually."
                )
        else:
            rising_fraction = 0.0

        diagnostics = detection.to_dict()
        diagnostics.update({
            "raw_points": len(df),
            "preload_method": "robust",
            "preload_found": detection.found,
            "preload_index": detection.index,
            "preload_force": detection.force,
            "preload_displacement": detection.displacement,
            "trimmed_start_index": detection.index,
            "trimmed_points_before_fracture_cut": len(x_trimmed),
            "analysis_points_after_fracture_cut": len(x_data),
            "analysis_source_start_index": int(analysis_source_indices[0]),
            "analysis_source_end_index": int(analysis_source_indices[-1]),
            "fracture_cut_source_index": int(retained_indices[fracture_cut_relative])
            if fracture_cut_relative is not None
            else None,
            "fracture_threshold_reached": fracture_cut_relative is not None,
            "postpeak_points_before_cut": postpeak_points,
            "postpeak_rising_fraction": rising_fraction,
            "fit_start_relative": best_start,
            "fit_end_relative": best_end,
            "fit_start_source_index": int(analysis_source_indices[best_start]),
            "fit_end_source_index": int(analysis_source_indices[best_end - 1]),
            "fit_r2": best_r2,
            "qc_status": qc_status,
            "qc_reasons": qc_reasons,
            "zero_reset": detection.zero_reset,
        })

        results_plot = {
            "x_data": x_data,
            "y_data": y_data,
            "a": a,
            "b": b,
            "best_start": best_start,
            "best_end": best_end,
            "yield_force_x": next_x,
            "yield_force_y": next_y,
            "raw_x_data": raw_x,
            "raw_y_data": raw_y,
            "preload_index": detection.index,
            "preload_threshold": preload,
            "analysis_source_indices": analysis_source_indices,
        }
        output_result = [
            max_value,
            a,
            next_y,
            postyield_displacement,
            auc,
        ]
        return {
            "results": output_result,
            "results_plot": results_plot,
            "diagnostics": diagnostics,
        }
    except Exception as e:
        logging.error(f"Error in robust analyse_data for {csv_file}: {e}")
        return None


def analyse_data(csv_file, x_column=DEFAULT_X_COLUMN, y_column=DEFAULT_Y_COLUMN,
                 min_window_size=10, max_window_size=20,
                 preload=DEFAULT_PRELOAD, YFC=None, dispc=None,
                 preload_mode=DEFAULT_PRELOAD_METHOD):
    """Dispatch to the traceable robust path or the preserved legacy path."""
    mode = str(preload_mode).strip().lower()
    if mode == "legacy":
        result = analyse_data_legacy(
            csv_file,
            x_column=x_column,
            y_column=y_column,
            min_window_size=min_window_size,
            max_window_size=max_window_size,
            preload=preload,
            YFC=YFC,
            dispc=dispc,
        )
        if result is not None:
            result["diagnostics"] = _legacy_diagnostics(
                csv_file, x_column, y_column, preload
            )
        return result
    if mode != "robust":
        logging.error(f"Unknown preload_mode: {preload_mode}")
        return None
    return analyse_data_robust(
        csv_file,
        x_column=x_column,
        y_column=y_column,
        min_window_size=min_window_size,
        max_window_size=max_window_size,
        preload=preload,
        YFC=YFC,
        dispc=dispc,
    )


def create_scatter_plot(title="3point", xlabel=DEFAULT_X_COLUMN, ylabel=DEFAULT_Y_COLUMN, output_image="scatter_plot.png", results_plot=None):
    """
        Creates a scatter plot.

        Args:
            title (str, optional): The title of the plot. Defaults to "3point".
            xlabel (str, optional): The label for the X-axis. Defaults to `DEFAULT_X_COLUMN`.
            ylabel (str, optional): The label for the Y-axis. Defaults to `DEFAULT_Y_COLUMN`.
            output_image (str, optional): The filename for the output image. Defaults to "scatter_plot.png".

    """
    try:
        x_data = results_plot["x_data"]
        y_data = results_plot["y_data"]

        plt.figure(figsize=(8, 6))
        plt.scatter(x_data, y_data, label="other")
        plt.title(title)
        plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.grid(True)

        best_start = results_plot["best_start"]
        best_end = results_plot["best_end"]

        a = results_plot['a']
        b = results_plot['b']

        x_fit_window = x_data[best_start:best_end]
        y_fit_window = y_data[best_start:best_end]

        plt.scatter(x_fit_window, y_fit_window, color='yellow', label="linear")
        x_fit_line = np.array([x_data.min(), x_data.max()])
        y_fit_line = a * x_fit_line + b
        plt.plot(x_fit_line, y_fit_line, color='red', label='linear line')

        plt.scatter(results_plot["yield_force_x"], results_plot["yield_force_y"], color='black', label="YF")
        plt.legend()
        plt.savefig(output_image)
        plt.close()

    except Exception as e:
        logging.error(f"Error in create_scatter_plot: {e}")
        plt.close()
        return None


def save_files(file_path, progress_callback, on_complete, min_window_size, max_window_size, failed_files_callback, preload=DEFAULT_PRELOAD,
               YFC=DEFAULT_Yield_Force_Constant, disp_c=DEFAULT_Displacement_Constant,
               preload_mode=DEFAULT_PRELOAD_METHOD):
    """
    Analyzes all CSV files in a specified folder and writes the results to an Excel file.

    Args:
        file_path (str): The path to the folder containing the CSV files.
        progress_callback (function): A callback function to update progress.
        on_complete (function): A callback function to call upon completion.
        min_window_size (int): The minimum size of the linear regression window.
        max_window_size (int): The maximum size of the linear regression window.
        failed_files_callback (function): A callback function to update the list of failed files.
        preload (float, optional): The preload value. Defaults to `DEFAULT_PRELOAD`.
    """
    all_results = []
    qc_rows = []
    png_dir = get_resource_path(file_path + OUTPUT_IMAGE_DIR)
    os.makedirs(png_dir, exist_ok=True)

    files = []
    failed_files = []

    for root, _, filenames in os.walk(file_path):
        for filename in filenames:
            if filename.endswith("Data.csv"):
                files.append(os.path.join(root, filename))

    total_files = len(files)

    for index, f in enumerate(files):
        try:
            file_name = os.path.basename(os.path.dirname(f))
            output_image = get_resource_path(os.path.join(png_dir, file_name + '.png'))
            analysis_output = analyse_data(f, x_column=DEFAULT_X_COLUMN, y_column=DEFAULT_Y_COLUMN, min_window_size=min_window_size,
                                           max_window_size=max_window_size, preload=preload, YFC=YFC, dispc=disp_c,
                                           preload_mode=preload_mode)
            if analysis_output is None:
                raise ValueError("analyse_data returned no result")
            result = analysis_output["results"]
            my_plot = analysis_output["results_plot"]
            diagnostics = analysis_output.get("diagnostics", {})
            create_scatter_plot(title=file_name, output_image=output_image, results_plot=my_plot)

            if result:
                all_results.append([file_name] + result)
                qc_rows.append([
                    file_name,
                    diagnostics.get("preload_method", preload_mode),
                    diagnostics.get("raw_points"),
                    diagnostics.get("threshold", preload),
                    diagnostics.get("last_baseline_index"),
                    diagnostics.get("preload_index"),
                    diagnostics.get("preload_force"),
                    diagnostics.get("preload_displacement"),
                    diagnostics.get("trimmed_points_before_fracture_cut"),
                    diagnostics.get("analysis_points_after_fracture_cut"),
                    diagnostics.get("zero_reset", False),
                    diagnostics.get("fit_start_source_index"),
                    diagnostics.get("fit_end_source_index"),
                    diagnostics.get("fracture_threshold_reached"),
                    diagnostics.get("qc_status", "WARNING"),
                    "; ".join(diagnostics.get("qc_reasons", [])),
                ])
            else:
                failed_files.append(os.path.basename(f))
        except Exception as e:
            logging.error(f"Error processing file {f}: {e}")
            failed_files.append(os.path.basename(f))
            qc_rows.append([
                os.path.basename(os.path.dirname(f)), preload_mode, None, preload,
                None, None, None, None, None, None, False, None, None, None,
                "FAIL", str(e),
            ])
        progress_callback(index + 1, total_files)

    wb = Workbook()
    ws = wb.active
    headers = ["File Name", "Max Force", "Stiffness", "Yield force", "Postyield Displacement", "Work to fracture"]
    ws.append(headers)
    last_first_char = None
    region_data = []
    for index, entry in enumerate(all_results):
        file_name = entry[0]
        match = re.match(r"([A-Za-z]+)", file_name)
        current_first_char = match.group(1) if match else ''
        if last_first_char is not None and current_first_char != last_first_char:
            if region_data:
                average_data = calculate_average(region_data)
                median_data = calculate_median(region_data)  # Call calculate_median
                ws.append(["Average"] + average_data)
                ws.append(["Median"] + median_data)  # Add median
                region_data = []
            ws.append([])
            ws.append([])
        region_data.append(entry[1:])
        ws.append(entry)
        last_first_char = current_first_char
        progress_callback(index + 1 + total_files, total_files * 2)

    # Handle the last region
    if region_data:
        average_data = calculate_average(region_data)
        median_data = calculate_median(region_data)  # Call calculate_median
        ws.append(["Average"] + average_data)
        ws.append(["Median"] + median_data)  # Add median

    ws.column_dimensions['A'].width = Excel_Type[0]
    ws.column_dimensions['B'].width = Excel_Type[1]
    ws.column_dimensions['C'].width = Excel_Type[2]
    ws.column_dimensions['D'].width = Excel_Type[3]
    ws.column_dimensions['E'].width = Excel_Type[4]
    ws.column_dimensions['F'].width = Excel_Type[5]

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '0.0000'

    qc_ws = wb.create_sheet("Preload QC")
    qc_headers = [
        "File Name", "Method", "Raw Points", "Preload Threshold",
        "Last Baseline Index", "Preload Index", "Preload Force",
        "Preload Displacement", "Points After Preload", "Analysis Points",
        "Zero Reset", "Fit Start Index", "Fit End Index",
        "50% Fracture Reached", "QC Status", "QC Reasons",
    ]
    qc_ws.append(qc_headers)
    for row in qc_rows:
        qc_ws.append(row)
    for row in qc_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
    for column in range(1, len(qc_headers) + 1):
        qc_ws.column_dimensions[chr(64 + column)].width = 18
    qc_ws.column_dimensions['P'].width = 70
    try:
        excel_file_path = get_resource_path(file_path + ".xlsx")
        wb.save(excel_file_path)
        logging.info(f"Excel file saved to {excel_file_path}")
    except Exception as e:
        logging.error(f"Error saving excel file: {e}")

    failed_files_callback(failed_files)
    on_complete()


def calculate_average(data):
    """Calculates the average of each column in the data."""
    num_cols = len(data[0])
    averages = []
    for i in range(num_cols):
        column_values = [float(row[i]) for row in data]
        avg = sum(column_values) / len(column_values)
        averages.append(avg)
    return averages


def calculate_median(data):
    """Calculates the median of each column of data"""
    num_cols = len(data[0])
    medians = []
    for i in range(num_cols):
        try:
            column_values = [float(row[i]) for row in data]
            median = statistics.median(column_values)
            medians.append(median)
        except (ValueError, TypeError) as e:
            logging.error(f"Error calculating median for column {i}: {e}")
            medians.append(None)  # Or use another appropriate default value
    return medians
